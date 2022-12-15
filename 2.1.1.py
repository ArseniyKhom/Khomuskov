import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side


class Vacancy:
    currency_value = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salary_from = self.salary_min_self(vacancy)
        self.salary_to = self.salary_max_self(vacancy)
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.salary_middle_self()
        self.area_name = vacancy['area_name']
        self.year = self.year_self(vacancy)

    def salary_min_self(self, vacancy):
        vacancy_result = vacancy['salary_from']
        float_result = float(vacancy_result)
        int_result = int(float_result)
        return int_result

    def salary_max_self(self, vacancy):
        vacancy_result = vacancy['salary_to']
        float_result = float(vacancy_result)
        int_result = int(float_result)
        return int_result

    def salary_middle_self(self):
        salary_currency = self.currency_value[self.salary_currency]
        salary_min = self.salary_from
        salary_max = self.salary_to
        middle_result = salary_currency * (salary_min + salary_max) / 2
        return middle_result

    def year_self(self, vacancy):
        year_result = vacancy['published_at'][:4]
        year_int_result = int(year_result)
        return year_int_result


class DataSet:
    def __init__(self, file_name, vacancy_name):
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    @staticmethod
    def increment(glossary, key_value, sum):
        if key_value in glossary:
            glossary[key_value] += sum
        else:
            glossary[key_value] = sum

    @staticmethod
    def average(glossary):
        new_glossary = {}
        for key_value, variables in glossary.items():
            new_glossary[key_value] = int(sum(variables) / len(variables))
        return new_glossary

    def csv_reader(self):
        open_file = open(self.file_name, encoding='utf-8-sig')
        with open_file:
            file_reader = csv.reader(open_file)
            title = next(file_reader)
            title_length = len(title)
            for line in file_reader:
                if '' not in line and len(line) == title_length:
                    yield dict(zip(title, line))

    def get_statistic(self):
        paycheck = {}
        paycheck_name = {}
        paycheck_city = {}
        number_available_seats = 0

        for vacancy_glossary in self.csv_reader():
            vacancy = Vacancy(vacancy_glossary)
            self.increment(paycheck, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.increment(paycheck_name, vacancy.year, [vacancy.salary_average])
            self.increment(paycheck_city, vacancy.area_name, [vacancy.salary_average])
            number_available_seats += 1

        vacancies_number = self.number_dict(paycheck)
        vacancies_name = self.number_dict(paycheck_name)

        if len(paycheck_name) == 0:
            paycheck_name = self.vacancy_dict(paycheck)
            vacancies_name = self.vacancies_dict(vacancies_number)

        statistics = self.average(paycheck)
        statistics2 = self.average(paycheck_name)
        statistics3 = self.average(paycheck_city)

        statistics4 = {}
        for year, salaries in paycheck_city.items():
            statistics4[year] = round(len(salaries) / number_available_seats, 4)
        statistics4 = self.statistics4_method(statistics4)
        statistics4.sort(key=lambda a: a[-1], reverse=True)
        statistics5 = statistics4.copy()
        statistics4 = dict(statistics4)
        statistics3 = list(filter(lambda a: a[0] in list(statistics4.keys()),
                                  [(key_value, value) for key_value, value in statistics3.items()]))
        statistics3.sort(key=lambda a: a[-1], reverse=True)
        statistics3 = dict(statistics3[:10])
        statistics5 = dict(statistics5[:10])

        return statistics, vacancies_number, statistics2, vacancies_name, statistics3, statistics5

    def number_dict(self, paycheck):
        result = dict([(key_value, len(value)) for key_value, value in paycheck.items()])
        return result

    def vacancy_dict(self, paycheck):
        for key_value, value in paycheck.items():
            return dict([(key_value, [0])])

    def vacancies_dict(self, vacancies_number):
        for key_value, value in vacancies_number.items():
            return dict([(key_value, 0)])

    def statistics4_method(self, statistics4):
        result = list(filter(lambda a: a[-1] >= 0.01, [(key_value, value) for key_value, value in statistics4.items()]))
        return result

    @staticmethod
    def print_statistic(statistics1, statistics2, statistics3, statistics4, statistics5, statistics6):
        print('Динамика уровня зарплат по годам: {0}'.format(statistics1))
        print('Динамика количества вакансий по годам: {0}'.format(statistics2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(statistics3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(statistics4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(statistics5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(statistics6))


class InputConnect:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        dataset = DataSet(self.file_name, self.vacancy_name)
        statistics1, statistics2, statistics3, statistics4, statistics5, statistics6 = dataset.get_statistic()
        dataset.print_statistic(statistics1, statistics2, statistics3, statistics4, statistics5, statistics6)

        self.report_method(statistics1, statistics2, statistics3, statistics4, statistics5, statistics6)

    def report_method(self, statistics1, statistics2, statistics3, statistics4, statistics5, statistics6):
        report = Report(self.vacancy_name, statistics1, statistics2, statistics3, statistics4, statistics5, statistics6)
        report.generate_excel()


class Report:
    def __init__(self, vacancy_name, statistics1, statistics2, statistics3, statistics4, statistics5, statistics6):
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.statistics1 = statistics1
        self.statistics2 = statistics2
        self.statistics3 = statistics3
        self.statistics4 = statistics4
        self.statistics5 = statistics5
        self.statistics6 = statistics6

    def generate_excel(self):
        line1 = self.wb.active
        line1.title = 'Статистика по годам'
        line1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])

        for annum in self.statistics1.keys(): line1.append([annum, self.statistics1[annum], self.statistics3[annum], self.statistics2[annum], self.statistics4[annum]])
        object_data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        pillar_widths = []

        pillar_widths = self.check_pillar(object_data, pillar_widths)

        for index, pillar_width in enumerate(pillar_widths, 1): line1.column_dimensions[get_column_letter(index)].width = pillar_width + 2
        object_data = []
        object_data.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])

        for (city1, value1), (city2, value2) in zip(self.statistics5.items(), self.statistics6.items()): object_data.append([city1, value1, '', city2, value2])
        line2 = self.wb.create_sheet('Статистика по городам')

        for line in object_data: line2.append(line)
        column_widths = []

        column_widths = self.check_column(column_widths, object_data)

        for index, column_width in enumerate(column_widths, 1): line2.column_dimensions[get_column_letter(index)].width = column_width + 2
        font_bold = Font(bold=True)

        for column in 'ABCDE':
            line1[column + '1'].font = font_bold
            line2[column + '1'].font = font_bold

        for index, _ in enumerate(self.statistics5): line2['E' + str(index + 2)].number_format = '0.00%'
        slim = Side(border_style='thin', color='00000000')

        for line in range(len(object_data)):
            for column in 'ABDE': line2[column + str(line + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)
        self.statistics1[1] = 1

        for line, _ in enumerate(self.statistics1):
            for column in 'ABCDE': line1[column + str(line + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)
        self.wb.save('report.xlsx')

    def check_column(self, column_widths, object_data):
        for line in object_data:
            for index, section in enumerate(line):
                section = str(section)
                if len(column_widths) < index or len(column_widths) == index:
                    column_widths += [len(section)]
                else:
                    if len(section) > column_widths[index]: column_widths[index] = len(section)
        return column_widths

    def check_pillar(self, object_data, pillar_widths):
        for line in object_data:
            for index, section in enumerate(line):
                if len(pillar_widths) < index or len(pillar_widths) == index:
                    pillar_widths += [len(section)]
                else:
                    if len(section) > pillar_widths[index]: pillar_widths[index] = len(section)
        return pillar_widths

InputConnect()